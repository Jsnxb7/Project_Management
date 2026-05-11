from flask import Blueprint, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from bson import ObjectId
from datetime import datetime, timezone
import csv
import io
import json
import os
import re
import sqlite3
import tempfile
import secrets
import time

from app import bcrypt
from database.db import users_collection, organizations_collection, teams_collection, projects_collection
from utils.response import ok, fail, warn
from utils.validators import valid_email, valid_password
from services.relation_service import sync_user_org_membership

portal_bp = Blueprint("portal_bp", __name__)
PORTAL_ROLES = ["Super User", "Admin", "Org Head", "Team Lead", "Member"]
BULK_IMPORT_SESSIONS = {}
BULK_IMPORT_SESSION_TTL_SECONDS = 60 * 60


def current_user():
    uid = get_jwt_identity()
    return users_collection.find_one({"_id": ObjectId(uid)}) if uid else None


def is_super_user(user):
    return user and user.get("portal_role", user.get("role")) == "Super User"


def oid(value):
    try:
        return ObjectId(value)
    except Exception:
        return None


def user_orgs(user_id):
    orgs = list(organizations_collection.find({
        "status": {"$ne": "deleted"},
        "members": {"$elemMatch": {"user_id": user_id, "status": "active"}}
    }, {"name": 1, "members": 1}).sort("name", 1))
    data = []
    for org in orgs:
        role = next((m.get("role") for m in org.get("members", []) if m.get("user_id") == user_id and m.get("status", "active") == "active"), "Member")
        data.append({"id": str(org["_id"]), "name": org.get("name"), "role": role})
    return data


def add_user_to_org(org_id, user_id, role="Member"):
    org_obj_id = oid(org_id)
    if not org_obj_id:
        return None, "Warning: choose a valid organization."
    org = organizations_collection.find_one({"_id": org_obj_id, "status": {"$ne": "deleted"}})
    if not org:
        return None, "Warning: selected organization was not found."
    if any(m.get("user_id") == user_id and m.get("status", "active") == "active" for m in org.get("members", [])):
        return org, "Warning: this user already belongs to the selected organization."
    sync_user_org_membership(org_obj_id, user_id, role, "active")
    return org, None


def _safe_int(value, default, minimum=1, maximum=100):
    try:
        number = int(value)
    except Exception:
        number = default
    return max(minimum, min(number, maximum))


def build_org_map_for_users(user_ids):
    """Fetch organization badges for only the visible page of users.

    This avoids the old N+1 pattern where every user card triggered its own
    organization query. It is the main optimization for the Portal Users page.
    """
    if not user_ids:
        return {}
    org_map = {str(uid): [] for uid in user_ids}
    orgs = organizations_collection.find({
        "status": {"$ne": "deleted"},
        "members.user_id": {"$in": user_ids},
    }, {"name": 1, "members.user_id": 1, "members.role": 1, "members.status": 1})
    for org in orgs:
        for member in org.get("members", []):
            uid = member.get("user_id")
            if uid in user_ids and member.get("status", "active") == "active":
                org_map.setdefault(str(uid), []).append({
                    "id": str(org["_id"]),
                    "name": org.get("name"),
                    "role": member.get("role", "Member"),
                })
    return org_map


def user_public(user, organizations=None, compact=False):
    orgs = organizations if organizations is not None else []
    data = {
        "id": str(user["_id"]),
        "name": user.get("name"),
        "email": user.get("email"),
        "portal_role": user.get("portal_role", user.get("role", "Member")),
        "role": user.get("portal_role", user.get("role", "Member")),
        "is_active": user.get("is_active", True),
        "organizations_count": len(orgs),
    }
    if not compact:
        data.update({
            "created_at": user.get("created_at").isoformat() if user.get("created_at") else None,
            "last_login": user.get("last_login").isoformat() if user.get("last_login") else None,
            "organizations": orgs,
            "relation_warnings": [] if orgs else ["This user is not assigned to any organization yet."],
        })
    return data


@portal_bp.get("/users")
@jwt_required()
def list_users():
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can view portal users.")

    q = (request.args.get("q") or "").strip()
    page = _safe_int(request.args.get("page"), 1, 1, 100000)
    limit = _safe_int(request.args.get("limit"), 25, 5, 100)
    compact = request.args.get("compact") == "true"

    query = {}
    if q:
        regex = {"$regex": re.escape(q), "$options": "i"}
        query = {"$or": [{"name": regex}, {"email": regex}, {"portal_role": regex}, {"role": regex}]}

    total_users = users_collection.count_documents({})
    filtered_users = users_collection.count_documents(query)
    active_users = users_collection.count_documents({"is_active": True})
    skip = (page - 1) * limit
    projection = {"name": 1, "email": 1, "portal_role": 1, "role": 1, "is_active": 1, "created_at": 1, "last_login": 1}
    visible_users = list(users_collection.find(query, projection).sort("created_at", -1).skip(skip).limit(limit))
    org_map = build_org_map_for_users([u["_id"] for u in visible_users]) if not compact else {}

    return ok("Portal users fetched", {
        "users": [user_public(u, org_map.get(str(u["_id"]), []), compact=compact) for u in visible_users],
        "meta": {
            "page": page,
            "limit": limit,
            "returned": len(visible_users),
            "filtered_users": filtered_users,
            "total_users": total_users,
            "active_users": active_users,
            "inactive_users": max(total_users - active_users, 0),
            "total_pages": max((filtered_users + limit - 1) // limit, 1),
            "query": q,
        }
    })


@portal_bp.post("/users")
@jwt_required()
def create_user():
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can create portal users.")

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    portal_role = data.get("portal_role") or data.get("role") or "Member"
    organization_id = data.get("organization_id")
    org_role = data.get("org_role") or "Member"

    if portal_role not in PORTAL_ROLES:
        return fail("Invalid portal role")
    if portal_role == "Super User" and users_collection.find_one({"portal_role": "Super User"}):
        return fail("A Super User already exists")
    if not name:
        return fail("Name is required")
    if not valid_email(email):
        return fail("Valid email is required")
    if not valid_password(password):
        return fail("Password must be at least 8 characters and contain letters and numbers")
    if users_collection.find_one({"email": email}):
        return fail("A user with this email already exists", 409)
    if org_role not in ["Admin", "Org Head", "Team Lead", "Member"]:
        return fail("Invalid organization role")

    now = datetime.now(timezone.utc)
    result = users_collection.insert_one({
        "name": name,
        "email": email,
        "password_hash": bcrypt.generate_password_hash(password).decode("utf-8"),
        "profile_image": None,
        "portal_role": portal_role,
        "role": portal_role,
        "is_active": True,
        "created_by": me["_id"],
        "created_at": now,
        "updated_at": now,
        "last_login": None,
    })
    warning_message = None
    if organization_id:
        _, warning_message = add_user_to_org(organization_id, result.inserted_id, org_role)
    if warning_message:
        return warn(f"Portal user was created, but organization assignment needs attention. {warning_message}", {"id": str(result.inserted_id)})
    return ok("Portal user created successfully", {"id": str(result.inserted_id)}, 201)


@portal_bp.patch("/users/<user_id>")
@jwt_required()
def update_user(user_id):
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can update portal users.")

    try:
        target_id = ObjectId(user_id)
    except Exception:
        return fail("Invalid user id")

    data = request.get_json() or {}
    update = {"updated_at": datetime.now(timezone.utc)}

    if "name" in data:
        name = (data.get("name") or "").strip()
        if not name:
            return fail("Name cannot be empty")
        update["name"] = name

    if "portal_role" in data or "role" in data:
        role = data.get("portal_role") or data.get("role")
        if role not in PORTAL_ROLES:
            return fail("Invalid portal role")
        if role == "Super User":
            existing = users_collection.find_one({"portal_role": "Super User", "_id": {"$ne": target_id}})
            if existing:
                return fail("Only one Super User is allowed")
        update["portal_role"] = role
        update["role"] = role

    if "is_active" in data:
        if target_id == me["_id"] and data.get("is_active") is False:
            return fail("You cannot deactivate yourself")
        update["is_active"] = bool(data.get("is_active"))

    result = users_collection.update_one({"_id": target_id}, {"$set": update})
    if result.matched_count == 0:
        return fail("User not found", 404)
    return ok("Portal user updated")


@portal_bp.post("/users/<user_id>/organizations")
@jwt_required()
def assign_user_to_org(user_id):
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can assign users from the portal user panel. Org Heads can add users from their organization page.")

    target_id = oid(user_id)
    if not target_id:
        return warn("Warning: choose a valid portal user.")
    target = users_collection.find_one({"_id": target_id, "is_active": True})
    if not target:
        return warn("Warning: portal user not found or inactive.")

    data = request.get_json() or {}
    role = data.get("role") or "Member"
    if role not in ["Admin", "Org Head", "Team Lead", "Member"]:
        return fail("Invalid organization role")

    org, warning_message = add_user_to_org(data.get("organization_id"), target_id, role)
    if warning_message:
        return warn(warning_message)
    return ok(f'{target.get("name", "User")} added to {org.get("name", "organization")}')


# ---------------- Bulk user import helpers ----------------
BULK_PORTAL_ROLES = ["Admin", "Org Head", "Team Lead", "Member"]
BULK_ORG_ROLES = ["Admin", "Org Head", "Team Lead", "Member"]
BULK_TEAM_ROLES = ["Team Lead", "Member"]
HEADER_ALIASES = {
    "name": ["name", "full_name", "fullname", "user_name", "username", "member_name"],
    "email": ["email", "email_address", "mail", "user_email", "member_email"],
    "password": ["password", "temp_password", "temporary_password", "default_password"],
    "portal_role": ["portal_role", "role", "user_role", "system_role"],
    "organization_id": ["organization_id", "org_id", "organisation_id", "company_id"],
    "organization_name": ["organization", "organization_name", "org", "org_name", "organisation", "organisation_name", "company"],
    "org_role": ["org_role", "organization_role", "organisation_role", "member_org_role"],
    "team_id": ["team_id"],
    "team_name": ["team", "team_name", "group", "group_name"],
    "team_role": ["team_role", "role_in_team"],
    "is_active": ["is_active", "active", "status"],
}


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_bool(value, default=True):
    raw = clean_cell(value).lower()
    if raw in ["", "none", "null"]:
        return default
    if raw in ["true", "1", "yes", "y", "active", "enabled"]:
        return True
    if raw in ["false", "0", "no", "n", "inactive", "disabled"]:
        return False
    return default


def read_json_rows(file_storage):
    payload = json.loads(file_storage.read().decode("utf-8-sig"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ["users", "members", "rows", "data"]:
            if isinstance(payload.get(key), list):
                return payload[key]
    raise ValueError("JSON must be a list or an object containing users/members/rows/data.")


def read_csv_rows(file_storage):
    text = file_storage.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def read_xlsx_rows(file_storage):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError("Excel import needs openpyxl. Add openpyxl to requirements.txt and redeploy.") from exc
    workbook = load_workbook(file_storage.stream, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_cell(h).lower() for h in rows[0]]
    data = []
    for raw in rows[1:]:
        item = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers)) if headers[i]}
        data.append(item)
    return data


def read_sqlite_rows(file_storage):
    suffix = os.path.splitext(file_storage.filename or "upload.db")[1] or ".db"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
        file_storage.save(temp.name)
        temp_path = temp.name
    try:
        conn = sqlite3.connect(temp_path)
        conn.row_factory = sqlite3.Row
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()]
        best_rows = []
        best_table = None
        for table in tables:
            safe_table = table.replace('"', '""')
            cols = [r[1].lower() for r in conn.execute(f'PRAGMA table_info("{safe_table}")').fetchall()]
            if any(alias in cols for alias in HEADER_ALIASES["email"]):
                best_rows = [dict(row) for row in conn.execute(f'SELECT * FROM "{safe_table}"').fetchall()]
                best_table = table
                break
        conn.close()
        if best_table is None:
            raise ValueError("No SQLite table with an email column was found.")
        return best_rows
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


def extract_rows(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".json"):
        return read_json_rows(file_storage)
    if filename.endswith(".csv"):
        return read_csv_rows(file_storage)
    if filename.endswith(".xlsx"):
        return read_xlsx_rows(file_storage)
    if filename.endswith((".db", ".sqlite", ".sqlite3")):
        return read_sqlite_rows(file_storage)
    if filename.endswith(".xls"):
        raise ValueError("Old .xls files are not supported directly. Save as .xlsx or CSV first.")
    raise ValueError("Unsupported file type. Upload JSON, CSV, XLSX, DB, SQLITE, or SQLITE3.")


def canonical_value(row, field):
    lowered = {str(k).strip().lower(): v for k, v in (row or {}).items()}
    for alias in HEADER_ALIASES[field]:
        if alias in lowered:
            return clean_cell(lowered.get(alias))
    return ""


def find_org(org_id=None, org_name=None):
    obj_id = oid(org_id) if org_id else None
    if obj_id:
        org = organizations_collection.find_one({"_id": obj_id, "status": {"$ne": "deleted"}})
        if org:
            return org
    name = clean_cell(org_name)
    if name:
        return organizations_collection.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}, "status": {"$ne": "deleted"}})
    return None


def find_team(team_id=None, team_name=None, org_id=None):
    obj_id = oid(team_id) if team_id else None
    query = {"status": {"$ne": "deleted"}}
    if org_id:
        query["organization_id"] = org_id
    if obj_id:
        team = teams_collection.find_one({"_id": obj_id, **query})
        if team:
            return team
    name = clean_cell(team_name)
    if name:
        query["name"] = {"$regex": f"^{re.escape(name)}$", "$options": "i"}
        return teams_collection.find_one(query)
    return None


def missing_org_key(name):
    return clean_cell(name).lower()


def missing_team_key(org_name, team_name):
    return f"{clean_cell(org_name).lower()}::{clean_cell(team_name).lower()}"


def normalize_import_row(row, index, defaults=None, seen_emails=None, mode="strict"):
    """Normalize one upload row.

    mode='analyze' keeps missing org/team references as conflicts instead of hard errors.
    mode='strict' blocks missing references unless commit explicitly auto-creates them first.
    """
    defaults = defaults or {}
    seen_emails = seen_emails if seen_emails is not None else set()
    name = canonical_value(row, "name") or clean_cell(row.get("name"))
    email = (canonical_value(row, "email") or clean_cell(row.get("email"))).lower()
    password = canonical_value(row, "password") or clean_cell(row.get("password")) or clean_cell(defaults.get("password"))
    portal_role = canonical_value(row, "portal_role") or clean_cell(row.get("portal_role")) or defaults.get("portal_role") or "Member"
    org_role = canonical_value(row, "org_role") or clean_cell(row.get("org_role")) or defaults.get("org_role") or "Member"
    team_role = canonical_value(row, "team_role") or clean_cell(row.get("team_role")) or defaults.get("team_role") or "Member"
    organization_id = canonical_value(row, "organization_id") or clean_cell(row.get("organization_id")) or defaults.get("organization_id") or ""
    organization_name = canonical_value(row, "organization_name") or clean_cell(row.get("organization_name")) or defaults.get("organization_name") or ""
    team_id = canonical_value(row, "team_id") or clean_cell(row.get("team_id")) or defaults.get("team_id") or ""
    team_name = canonical_value(row, "team_name") or clean_cell(row.get("team_name")) or defaults.get("team_name") or ""
    is_active = normalize_bool(canonical_value(row, "is_active") or row.get("is_active"), True)

    errors = []
    warnings = []
    conflicts = []
    if not name:
        errors.append("Name is required.")
    if not valid_email(email):
        errors.append("Valid email is required.")
    if email and email in seen_emails:
        errors.append("Duplicate email inside upload.")
    if email:
        seen_emails.add(email)
    if email and users_collection.find_one({"email": email}):
        errors.append("Email already exists in portal.")
    if not valid_password(password):
        errors.append("Password must be at least 8 characters and contain letters and numbers.")
    if portal_role not in BULK_PORTAL_ROLES:
        errors.append("Portal role must be Admin, Org Head, Team Lead, or Member. Bulk import cannot create Super User accounts.")
    if org_role not in BULK_ORG_ROLES:
        errors.append("Organization role must be Admin, Org Head, Team Lead, or Member.")
    if team_role not in BULK_TEAM_ROLES:
        errors.append("Team role must be Team Lead or Member.")

    org = find_org(organization_id, organization_name)
    team = None
    if organization_id or organization_name:
        if not org:
            msg = "Organization was provided but not found."
            if mode == "analyze" and organization_name:
                conflicts.append(msg)
            else:
                errors.append(msg)
    elif defaults.get("require_org"):
        warnings.append("No organization selected. User will be created without org membership.")

    if team_id or team_name:
        team = find_team(team_id, team_name, org.get("_id") if org else None)
        if not team:
            msg = "Team was provided but not found inside the selected organization."
            if mode == "analyze" and team_name:
                conflicts.append(msg)
            else:
                errors.append(msg)
        elif org and team.get("organization_id") != org.get("_id"):
            errors.append("Selected team does not belong to selected organization.")
        elif not org:
            org = organizations_collection.find_one({"_id": team.get("organization_id")})
            organization_id = str(org["_id"]) if org else ""
            organization_name = org.get("name") if org else ""
    elif defaults.get("require_team"):
        warnings.append("No team selected. User will not be added to a team.")

    return {
        "row": index,
        "name": name,
        "email": email,
        "password": password,
        "portal_role": portal_role,
        "organization_id": str(org["_id"]) if org else "",
        "organization_name": org.get("name") if org else organization_name,
        "org_role": org_role,
        "team_id": str(team["_id"]) if team else "",
        "team_name": team.get("name") if team else team_name,
        "team_role": team_role,
        "is_active": is_active,
        "errors": errors,
        "warnings": warnings,
        "conflicts": conflicts,
        "valid": len(errors) == 0,
        "importable": len(errors) == 0 and len(conflicts) == 0,
    }


def create_import_org(name, actor_id):
    name = clean_cell(name)
    if not name:
        return None
    existing = find_org(org_name=name)
    if existing:
        return existing
    now = datetime.now(timezone.utc)
    result = organizations_collection.insert_one({
        "name": name,
        "description": "Created automatically during bulk user import.",
        "visibility": "Private",
        "status": "active",
        "created_by": actor_id,
        "members": [{"user_id": actor_id, "role": "Admin", "status": "active", "joined_at": now}],
        "created_at": now,
        "updated_at": now,
        "created_from_bulk_import": True,
    })
    sync_user_org_membership(result.inserted_id, actor_id, "Admin", "active", actor_id)
    return organizations_collection.find_one({"_id": result.inserted_id})


def create_import_team(name, org_id, actor_id):
    name = clean_cell(name)
    org_obj_id = oid(org_id)
    if not name or not org_obj_id:
        return None
    existing = find_team(team_name=name, org_id=org_obj_id)
    if existing:
        return existing
    now = datetime.now(timezone.utc)
    result = teams_collection.insert_one({
        "organization_id": org_obj_id,
        "name": name,
        "description": "Created automatically during bulk user import.",
        "team_lead_id": None,
        "members": [],
        "assigned_project_ids": [],
        "status": "active",
        "created_by": actor_id,
        "created_at": now,
        "updated_at": now,
        "created_from_bulk_import": True,
    })
    return teams_collection.find_one({"_id": result.inserted_id})


def resolve_missing_references(rows, actor_id, create_orgs=False, create_teams=False):
    created_orgs = []
    created_teams = []
    org_cache = {}
    team_cache = {}

    if create_orgs:
        for row in rows:
            if row.get("organization_id") or not row.get("organization_name"):
                continue
            key = missing_org_key(row.get("organization_name"))
            if key not in org_cache:
                org = create_import_org(row.get("organization_name"), actor_id)
                if org:
                    org_cache[key] = org
                    created_orgs.append({"id": str(org["_id"]), "name": org.get("name")})
            if key in org_cache:
                row["organization_id"] = str(org_cache[key]["_id"])
                row["organization_name"] = org_cache[key].get("name")

    if create_teams:
        for row in rows:
            if row.get("team_id") or not row.get("team_name"):
                continue
            org_id = row.get("organization_id")
            if not org_id and row.get("organization_name") and create_orgs:
                org = find_org(org_name=row.get("organization_name"))
                if org:
                    org_id = str(org["_id"])
                    row["organization_id"] = org_id
                    row["organization_name"] = org.get("name")
            if not org_id:
                continue
            key = missing_team_key(row.get("organization_name") or org_id, row.get("team_name"))
            if key not in team_cache:
                team = create_import_team(row.get("team_name"), org_id, actor_id)
                if team:
                    team_cache[key] = team
                    created_teams.append({"id": str(team["_id"]), "name": team.get("name"), "organization_id": str(team.get("organization_id"))})
            if key in team_cache:
                row["team_id"] = str(team_cache[key]["_id"])
                row["team_name"] = team_cache[key].get("name")
    return created_orgs, created_teams


def summarize_import_rows(rows):
    missing_orgs = {}
    missing_teams = {}
    duplicate_upload_emails = []
    existing_emails = []
    invalid_rows = []
    warning_rows = []
    for row in rows:
        for error in row.get("errors", []):
            if "Duplicate email" in error:
                duplicate_upload_emails.append({"row": row["row"], "email": row.get("email")})
            elif "already exists" in error:
                existing_emails.append({"row": row["row"], "email": row.get("email")})
        if row.get("organization_name") and not row.get("organization_id"):
            key = missing_org_key(row["organization_name"])
            missing_orgs.setdefault(key, {"name": row["organization_name"], "rows": []})["rows"].append(row["row"])
        if row.get("team_name") and not row.get("team_id"):
            org_label = row.get("organization_name") or "No organization"
            key = missing_team_key(org_label, row["team_name"])
            missing_teams.setdefault(key, {"name": row["team_name"], "organization": org_label, "rows": []})["rows"].append(row["row"])
        if row.get("errors"):
            invalid_rows.append({"row": row["row"], "email": row.get("email"), "errors": row.get("errors")})
        elif row.get("warnings"):
            warning_rows.append({"row": row["row"], "email": row.get("email"), "warnings": row.get("warnings")})

    return {
        "missing_organizations": list(missing_orgs.values()),
        "missing_teams": list(missing_teams.values()),
        "duplicate_upload_emails": duplicate_upload_emails,
        "existing_emails": existing_emails,
        "invalid_rows": invalid_rows[:50],
        "invalid_rows_total": len(invalid_rows),
        "warning_rows": warning_rows[:50],
        "warning_rows_total": len(warning_rows),
    }


def build_import_stats(rows):
    return {
        "total_rows": len(rows),
        "ready_rows": sum(1 for r in rows if r.get("importable")),
        "conflict_rows": sum(1 for r in rows if r.get("valid") and r.get("conflicts")),
        "invalid_rows": sum(1 for r in rows if r.get("errors")),
        "warnings": sum(len(r.get("warnings", [])) for r in rows),
        "with_organization": sum(1 for r in rows if r.get("organization_id") or r.get("organization_name")),
        "with_team": sum(1 for r in rows if r.get("team_id") or r.get("team_name")),
    }


def add_user_to_team(team_id, user_id, team_role, actor_id):
    team_obj_id = oid(team_id)
    if not team_obj_id:
        return False, "Invalid team id."
    team = teams_collection.find_one({"_id": team_obj_id, "status": "active"})
    if not team:
        return False, "Team not found."
    now = datetime.now(timezone.utc)
    role = "Team Lead" if team_role == "Team Lead" else "Member"
    if role == "Team Lead":
        teams_collection.update_one({"_id": team_obj_id}, {"$set": {"team_lead_id": user_id, "updated_at": now}})
    if any(m.get("user_id") == user_id for m in team.get("members", [])):
        teams_collection.update_one({"_id": team_obj_id, "members.user_id": user_id}, {"$set": {"members.$.role": role, "members.$.status": "active", "members.$.updated_at": now, "updated_at": now}})
    else:
        teams_collection.update_one({"_id": team_obj_id}, {"$push": {"members": {"user_id": user_id, "role": role, "status": "active", "joined_at": now, "added_by": actor_id}}, "$set": {"updated_at": now}})
    for project_id in team.get("assigned_project_ids", []):
        project = projects_collection.find_one({"_id": project_id, "status": "active"})
        if not project:
            continue
        existing = next((m for m in project.get("members", []) if m.get("user_id") == user_id), None)
        if existing:
            projects_collection.update_one({"_id": project_id, "members.user_id": user_id}, {"$set": {"members.$.status": "active", "members.$.updated_at": now, "updated_at": now}, "$addToSet": {"members.$.source_team_ids": team_obj_id}})
        else:
            projects_collection.update_one({"_id": project_id}, {"$push": {"members": {"user_id": user_id, "role": role, "status": "active", "joined_at": now, "added_by": actor_id, "source_team_ids": [team_obj_id], "added_by_team_only": True}}, "$set": {"updated_at": now}})
    return True, None


@portal_bp.get("/import/options")
@jwt_required()
def bulk_import_options():
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can use bulk import.")
    orgs = list(organizations_collection.find({"status": {"$ne": "deleted"}}, {"name": 1}).sort("name", 1))
    teams = list(teams_collection.find({"status": "active"}, {"name": 1, "organization_id": 1}).sort("name", 1))
    org_name_by_id = {str(o["_id"]): o.get("name") for o in orgs}
    return ok("Bulk import options fetched", {
        "organizations": [{"id": str(o["_id"]), "name": o.get("name")} for o in orgs],
        "teams": [{"id": str(t["_id"]), "name": t.get("name"), "organization_id": str(t.get("organization_id")), "organization_name": org_name_by_id.get(str(t.get("organization_id")), "")} for t in teams],
        "required_columns": ["name", "email"],
        "optional_columns": ["password", "portal_role", "organization_id", "organization_name", "org_role", "team_id", "team_name", "team_role", "is_active"],
        "supported_files": [".json", ".csv", ".xlsx", ".db", ".sqlite", ".sqlite3"],
    })


@portal_bp.post("/import/preview")
@jwt_required()
def preview_bulk_import():
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can preview bulk user imports.")
    upload = request.files.get("file")
    if not upload:
        return fail("Upload a JSON, CSV, XLSX, or SQLite DB file.")
    defaults = {
        "password": request.form.get("default_password") or "",
        "portal_role": request.form.get("default_portal_role") or "Member",
        "organization_id": request.form.get("default_organization_id") or "",
        "org_role": request.form.get("default_org_role") or "Member",
        "team_id": request.form.get("default_team_id") or "",
        "team_role": request.form.get("default_team_role") or "Member",
        "require_org": request.form.get("require_org") == "true",
        "require_team": request.form.get("require_team") == "true",
    }
    try:
        raw_rows = extract_rows(upload)
    except Exception as exc:
        return fail(str(exc))
    if not raw_rows:
        return fail("No rows were found in the uploaded file.")
    if len(raw_rows) > 1000:
        return fail("Please import at most 1000 users at a time.")
    seen = set()
    rows = [normalize_import_row(row, idx + 1, defaults, seen, mode="analyze") for idx, row in enumerate(raw_rows)]
    stats = build_import_stats(rows)
    issues = summarize_import_rows(rows)

    # Keep the analyzed rows on the server instead of sending 1000 full row objects
    # with passwords back to the browser. The browser receives counts, conflicts,
    # and an import token only. This makes the page much faster and safer.
    now_ts = time.time()
    for token, session in list(BULK_IMPORT_SESSIONS.items()):
        if now_ts - session.get("created_at", 0) > BULK_IMPORT_SESSION_TTL_SECONDS:
            BULK_IMPORT_SESSIONS.pop(token, None)
    import_token = secrets.token_urlsafe(24)
    BULK_IMPORT_SESSIONS[import_token] = {
        "rows": rows,
        "created_at": now_ts,
        "actor_id": str(me["_id"]),
    }
    sample_rows = [{k: v for k, v in row.items() if k != "password"} for row in rows[:8]]
    return ok("File analyzed. Review only conflicts and invalid rows below.", {
        "import_token": import_token,
        "row_count": len(rows),
        "sample_rows": sample_rows,
        "stats": stats,
        "issues": issues,
    })


@portal_bp.post("/import/commit")
@jwt_required()
def commit_bulk_import():
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can import users.")
    data = request.get_json() or {}
    submitted_rows = data.get("rows") or []
    import_token = data.get("import_token")
    if not submitted_rows and import_token:
        session = BULK_IMPORT_SESSIONS.get(import_token)
        if not session or session.get("actor_id") != str(me["_id"]):
            return fail("Import session expired. Re-analyze the file and import again.")
        if time.time() - session.get("created_at", 0) > BULK_IMPORT_SESSION_TTL_SECONDS:
            BULK_IMPORT_SESSIONS.pop(import_token, None)
            return fail("Import session expired. Re-analyze the file and import again.")
        submitted_rows = session.get("rows") or []
    create_missing_orgs = bool(data.get("create_missing_orgs"))
    create_missing_teams = bool(data.get("create_missing_teams"))
    skip_existing = data.get("skip_existing", True)
    skip_invalid = data.get("skip_invalid", True)
    if not submitted_rows:
        return fail("No analyzed rows were found for import.")
    if len(submitted_rows) > 1000:
        return fail("Please import at most 1000 users at a time.")

    working_rows = [dict(r) for r in submitted_rows]
    created_orgs, created_teams = resolve_missing_references(working_rows, me["_id"], create_missing_orgs, create_missing_teams)

    seen = set()
    normalized = [normalize_import_row(row, idx + 1, {}, seen, mode="strict") for idx, row in enumerate(working_rows)]
    invalid_rows = [r for r in normalized if r.get("errors")]
    if invalid_rows and not skip_invalid:
        return warn("Some rows are invalid. Turn on skip invalid rows or fix the file.", {"stats": build_import_stats(normalized), "issues": summarize_import_rows(normalized)})

    created = 0
    org_assigned = 0
    team_assigned = 0
    skipped = []
    now = datetime.now(timezone.utc)
    for row in normalized:
        if row.get("errors"):
            skipped.append({"row": row["row"], "email": row.get("email"), "reason": "; ".join(row.get("errors", []))})
            continue
        if users_collection.find_one({"email": row["email"]}):
            if skip_existing:
                skipped.append({"row": row["row"], "email": row.get("email"), "reason": "Email already exists."})
                continue
            return fail(f'Email already exists: {row["email"]}', 409)
        result = users_collection.insert_one({
            "name": row["name"],
            "email": row["email"],
            "password_hash": bcrypt.generate_password_hash(row["password"]).decode("utf-8"),
            "profile_image": None,
            "portal_role": row["portal_role"],
            "role": row["portal_role"],
            "is_active": bool(row.get("is_active", True)),
            "created_by": me["_id"],
            "created_at": now,
            "updated_at": now,
            "last_login": None,
            "imported_by_bulk": True,
        })
        created += 1
        user_id = result.inserted_id
        if row.get("organization_id"):
            sync_user_org_membership(row["organization_id"], user_id, row.get("org_role") or "Member", "active", me["_id"])
            org_assigned += 1
        if row.get("team_id"):
            ok_team, team_error = add_user_to_team(row["team_id"], user_id, row.get("team_role") or "Member", me["_id"])
            if ok_team:
                team_assigned += 1
            else:
                skipped.append({"row": row["row"], "email": row["email"], "reason": f"User created, but team assignment failed: {team_error}"})
    if import_token:
        BULK_IMPORT_SESSIONS.pop(import_token, None)
    return ok("Bulk import completed", {
        "created": created,
        "organization_assignments": org_assigned,
        "team_assignments": team_assigned,
        "created_organizations": created_orgs,
        "created_teams": created_teams,
        "skipped": skipped,
        "stats": {"submitted": len(submitted_rows), "created": created, "skipped": len(skipped), "created_organizations": len(created_orgs), "created_teams": len(created_teams)},
    }, 201)


@portal_bp.get("/summary")
@jwt_required()
def portal_summary():
    me = current_user()
    if not is_super_user(me):
        return warn("Warning: only the Super User can view portal summary.")
    return ok("Portal summary fetched", {
        "total_users": users_collection.count_documents({}),
        "active_users": users_collection.count_documents({"is_active": True}),
        "organizations": organizations_collection.count_documents({"status": {"$ne": "deleted"}}),
    })
